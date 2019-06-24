# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

# import pymongo
# from kuaididanhao import settings
# from kuaididanhao.items import KuaididanhaoItem
# from openpyxl import workbook
import datetime
from kuaididanhao import settings
from openpyxl import load_workbook
from kuaididanhao.util.globalLog import ta_log


class KuaididanhaoPipeline(object):
    # 保存到Excel
    file_home = r''+settings.excel_path
    wb = load_workbook(filename=file_home)
    try:
        ws = wb['划款审单下载']
    except (Exception):
        ws = wb['Sheet1']
    # 新增列
    ws["R2"] = '发货单号'
    ws["S2"] = '物流状态'
    ws["T2"] = '发货地址'
    ws["U2"] = '当前地址'
    ws["V2"] = '当前时间'
    ws["W2"] = '事件'
    rowlenth = ws.max_row - 2
    s = 0
    start = datetime.datetime.now()

    def process_item(self, item, spider):
        self.s += 1
        end = datetime.datetime.now()
        d = (end - self.start).seconds
        read_len = spider.read_len

        trackNo = item['trackNo']
        trackNo = str(trackNo).strip()
        trackStatus = item['trackStatus']
        if item['trackOrigin']==None:
            trackOrigin = '未知'
        else:
            trackOrigin = item['trackOrigin']

        if item['trackTarget']==None:
            trackTarget = '未知'
        else:
            trackTarget = item['trackTarget']

        if item['trackTime']==None:
            trackTime = '未知'
        else:
            trackTime = item['trackTime']

        if item['trackEvent']==None:
            trackEvent = '未知'
        else:
            trackEvent = item['trackEvent']
        i = 0
        try:
            for cell in list(self.ws.columns)[10]:  # 遍历第11列
                i += 1
                if trackNo in str(cell.value).upper():
                    self.ws["R" + str(i)] = trackNo
                    self.ws["S" + str(i)] = trackStatus
                    self.ws["T" + str(i)] = trackOrigin
                    self.ws["U" + str(i)] = trackTarget
                    self.ws["V" + str(i)] = trackTime
                    self.ws["W" + str(i)] = trackEvent
                    ta_log.info('解析第 %s 条数据,总耗时 %d 秒' % (self.s, d))
                    break
            if self.s % 200 == 0:
                ta_log.info('excel总条数:'+self.rowlenth)
                ta_log.info('已保存第 %s 条数据,总耗时 %d 秒' % (self.s, d))
                self.wb.save(self.file_home)
            if self.rowlenth == (self.s+1) or read_len == (self.s+1):
                ta_log.info('excel总条数:'+self.rowlenth)
                ta_log.info('已保存第 %s 条数据(最后一条),总耗时 %d 秒' % (self.s, d))
                self.wb.save(self.file_home)
                exit(0)
            return item
        except BaseException:
            self.wb.close()
            ta_log.error('保存excel报错')
    # 保存到MongoDB
    # def __init__(self):
    #     host = settings.mongo_host
    #     port = settings.mongo_port
    #     dbname = settings.mongo_db_name
    #     sheetname = settings.mongo_db_collection
    #     client = pymongo.mongo_client.MongoClient(host=host, port=port)
    #     mydb = client[dbname]
    #     self.post = mydb[sheetname]
    #
    # def process_item(self, item, spider):
    #     data = dict(item)
    #     self.post.insert(data)
    #     return item
