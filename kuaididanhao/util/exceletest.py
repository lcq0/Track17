# import xlwt
# from datetime import date, datetime
# from random import random
# from openpyxl import workbook
# from openpyxl.utils import get_column_letter
import xlrd
from collections import deque
# class Excel:
#     def read_excel(self):
#         # 打开文件
#         workbook = xlrd.open_workbook(r'D:\PC_WorkSpace\back\82045002.xls')
#         print(workbook.sheet_names())
#         sheet1_name = workbook.sheet_names()[0]
#         # 根据索引或者名称获取sheet内容
#         sheet1 = workbook.sheet_by_index(0)
#         # sheet1 = workbook.sheet_by_name('sheet1')
#         # sheet的名称，行数，列数
#         print(sheet1.name, sheet1.nrows, sheet1.ncols)
#
#         # 获取整行或整列的值（数组）
#         # rows = sheet1.row_values(2)  # 获取第3行内容
#         cols3 = sheet1.col_values(2)  # 获取第3列内容
#         cols11 = sheet1.col_values(10)  # 获取第11列内容
#         i = 0
#         queue_col3 = deque(cols3)
#         queue_col11 = deque(cols11)
#         while(i<2):
#             i += 1
#             queue_col3.popleft()
#             queue_col11.popleft()
#         # 获取单元格内容
#         track_url = sheet1.cell(2, 12)
#         queue_col11.appendleft(track_url)
#         # print(queue_col11)
#         # print(queue_col11)
#         # print(sheet1.cell(0, 0))
#         # print(sheet1.cell_value(0, 0))
#         # print(sheet1.row(0)[0])
#         return queue_col11
#
#     t_queue = read_excel("main")
#     i = 0
#     t = t_queue.popleft()
#     if isinstance(t, float):
#         track_no = '#nums=' + str(int(t))
#     else:
#         track_no = '#nums=' + str(t)
#     while(i<39):
#         i += 1
#         t = t_queue.popleft()
#         if isinstance(t, float):
#             track_no = track_no + ',' + str(int(t))
#         else:
#             track_no = track_no + ',' + str(t)
#     print(track_no)

from openpyxl import load_workbook
import datetime
from bs4 import BeautifulSoup
import requests


class Excel2:
    i = 0
    j = 0
    ip_list = []
    index = 0
    user_file = open('C:\TrackIP\IPList.txt', 'r')
    lines = user_file.readlines()
    for line in lines:
        ip_list.insert(index, line.strip())
        index += 1
    def prindd(self):
        while True:
            if len(self.ip_list) <= 0:
                print("已经没有可用的IP地址了，请更换IP并重启程序！")
            if len(self.ip_list) - 1 > self.j:
                temp_ip = self.ip_list[self.j]
            else:
                self.j = 0
                temp_ip = self.ip_list[self.j]
            self.j = self.j + 1
            if len(temp_ip) < 1:
                print('全部IP地址都已无效，请重新开始')
            else:
                try:
                    requests.get('https://t.17track.net', proxies={"http": temp_ip})
                except:
                    print("删除无效代理IP=" + self.ip_list.pop(self.j - 1))
                    self.j = self.j - 1
                    continue
                else:
                    print("有效的代理IP=" + temp_ip)
                    continue
        print('请求代理IP：', temp_ip)
    prindd('e')

    # try:
    #     url = 'http://ip.tool.chinaz.com/'
    #     ip = '115.153.10.21:42429'
    #     proxies = {
    #         'http': ip,
    #     }
    #     r = requests.get(url, proxies=proxies)
    #     soup = BeautifulSoup(r.text, 'lxml')
    #     parent_node = soup.find(class_="IpMRig-tit")
    #     for i in parent_node.find_all('dd'):
    #         print(i.get_text())
    # except BaseException:
    #     print('无效IP地址')

    # def fibonacci(n):
    #     if n == 0:
    #         return 0
    #     elif n == 1:
    #         return 1
    #     else:
    #         return fibonacci(n - 1) + fibonacci(n - 2)
    #
    # print([fibonacci(x) for x in range(10)])


    # def IPcheck(self, ip):
    #     print(ip)
    #     '''代理IP地址（高匿）'''
    #     '''head 信息'''
    #     head = {
    #         'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36',
    #         'Connection': 'keep-alive'}
    #     '''http://icanhazip.com会返回当前的IP地址'''
    #     p = requests.get('http://icanhazip.com', headers=head, proxies=ip)
    #     print(p.text)
    # temp_ip = '49.68.157.57:4243'
    # IPcheck("main", temp_ip)

    def update_excel(self):
        file_home = r'D:\PC_WorkSpace\back\82067--5229笔1.xlsx'
        wb = load_workbook(filename=file_home)
        sheet_ranges = wb['Sheet1']
        print(sheet_ranges['A1'].value)
        ws = wb['Sheet1']
        # 新增列
        ws["R2"] = '发货单号'
        ws["S2"] = '物流状态'
        ws["T2"] = '发货地址'
        ws["U2"] = '当前地址'
        ws["V2"] = '当前时间'
        ws["W2"] = '事件'
        rowlenth = ws.max_row

        starttime = datetime.datetime.now()

        i = 0
        for cell in list(ws.columns)[0]:  # 遍历第11列
            i += 1
            if 'RV880239947CN' in str(cell.value):
                ws["K" + str(i)] = 'LF003262819SG'
                ws["L" + str(i)] = 'Delivery'
                ws["M" + str(i)] = 'China'
                ws["N" + str(i)] = 'American'
                ws["O" + str(i)] = '早上一点'
                ws["P" + str(i)] = '已成功收货'
                print('调用save！')
                # wb.save(file_home)
                break
        j = 0
        while j < rowlenth:
            j += 1
            if rowlenth % j == 0 and j != 1:
                print('保存excel！', j)
            if rowlenth == j:
                wb.save(file_home)
                print('最后一次保存excel')
        endtime = datetime.datetime.now()

        print((endtime - starttime).seconds)

    # update_excel("main")